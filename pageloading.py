#! python3
# pagetime.py

# Script to get page loading time

import time
import sys  
import openpyxl
import os
from selenium import webdriver
from statistics import mean
import collections

wb = openpyxl.load_workbook(os.path.join(sys.path[0], 'Times.xlsx'))
sheet = wb.get_sheet_by_name("Times")
driver = webdriver.PhantomJS()

elapsed_time =[]
newRow = sheet.max_row +1

websites= (
('Website 1','https://www.google.ca'), 
('Website 2','https://www.google.ca'),
('Website 3','https://www.google.ca'),
('Website 4','https://www.google.ca'), 
('Website 5','https://www.google.ca'),
('Website 6','https://www.google.ca'),
('Website 7','https://www.google.ca'),
('Website 8','https://www.google.ca'),
('Website 9','https://www.google.ca'),
('Website 10','https://www.google.ca')
)
websites = collections.OrderedDict(websites)
	
def main():
	print('Getting page load time\n')
	get_times(websites)
	print('\nWriting to excel sheet')
	write_to_sheet()
	driver.quit()

def get_times(website_dictionary):
	"""Goes through the dictionary and gets the load time, then appends the result to list"""
	for key, value in website_dictionary.items():
		average_time = []
		print(key,'... Started')
		"""Gets the average of 5 measurements - first one is skipped because it is usually something unusually high"""
		for i in range(0,6):
			start = time.time()
			driver.get(value)
			end = time.time()
			calculate = end - start
			if i == 0:
				continue
			print(calculate)
			average_time.append(calculate)
		print('Average is: '+ str(mean(average_time)))
		elapsed_time.append(mean(average_time))
		print(key,' ... Done\n')
		del average_time[:]

def write_to_sheet():
	"""writes the result to the excel spreasheet"""
	sheet.cell(row = newRow, column = 1).value = time.strftime("%Y-%m-%d %H:%M")
	for colNum in range(2,sheet.max_column+1):
		sheet.cell(row = newRow, column = colNum).value = elapsed_time[colNum-2]
	wb.save('Times.xlsx')

if __name__ == '__main__':
	sys.exit(main())

